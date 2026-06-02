# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  INFORME DIARIO REGIONAL — Comfenalco Antioquia                             ║
# ║  100% gratuito: Groq + Tavily + Streamlit Cloud                             ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

import streamlit as st
from tavily import TavilyClient
from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from email.utils import parsedate_to_datetime
import json, re, io, time, unicodedata

st.set_page_config(
    page_title="Informe Diario Regional | Comfenalco",
    page_icon="https://i.imgur.com/RFdkaOo.png",
    layout="wide",
)

# ══════════════════════════════════════════════════════════════════════════════
#  DATOS
# ══════════════════════════════════════════════════════════════════════════════
SUBREGIONES = {
    "BAJO CAUCA":         ["CACERES", "CAUCASIA", "EL BAGRE", "NECHI", "TARAZÁ", "ZARAGOZA"],
    "MAGDALENA MEDIO":    ["CARACOLI", "MACEO", "PUERTO BERRIO", "PUERTO NARE", "PUERTO TRIUNFO", "YONDO"],
    "NORDESTE":           ["AMALFI", "ANORI", "CISNEROS", "REMEDIOS", "SAN ROQUE", "SANTO DOMINGO", "SEGOVIA", "VEGACHI", "YALI", "YOLOMBO"],
    "NORTE":              ["ANGOSTURA", "BELMIRA", "BRICEÑO", "CAMPAMENTO", "CAROLINA DEL PRINCIPE", "DON MATIAS", "ENTRERRIOS", "GOMEZ PLATA", "GUADALUPE", "ITUANGO", "SAN ANDRES DE CUERQUIA", "SANTA ROSA DE OSOS", "TOLEDO", "VALDIVIA", "YARUMAL"],
    "OCCIDENTE":          ["ABRIAQUI", "ANZA", "BURITICA", "CAICEDO", "CAÑAS GORDAS", "DABEIBA", "EBEJICO", "FRONTINO", "HELICONIA", "LIBORINA", "OLAYA", "PEQUE", "SABANALARGA", "SAN JERONIMO", "SANTA FE DE ANTIOQUIA", "SOPETRÁN", "URAMITA"],
    "ORIENTE":            ["ABEJORRAL", "ALEJANDRIA", "ARGELIA", "COCORNÁ", "EL CARMEN DE VIBORAL", "EL PEÑOL", "EL RETIRO", "EL SANTUARIO", "GRANADA", "GUARNE", "GUATAPE", "LA CEJA", "LA UNION", "MARINILLA", "RIONEGRO", "SAN CARLOS", "SAN FRANCISCO", "SAN LUIS", "SAN RAFAEL", "SAN VICENTE", "SONSON"],
    "URABÁ":              ["APARTADÓ", "ARBOLETES", "CAREPA", "CHIGORODÓ", "MURINDO", "MUTATA", "NECOCLI", "SAN JUAN DE URABA", "SAN PEDRO DE URABA", "TURBO", "VIGIA DEL FUERTE"],
    "SUROESTE":           ["AMAGA", "ANDES", "ANGELOPOLIS", "BETANIA", "BETULIA", "CARAMANTA", "CIUDAD BOLIVAR", "CONCORDIA", "FREDONIA", "HISPANIA", "JARDIN", "JERICO", "LA PINTADA", "MONTEBELLO", "PUEBLORICO", "SALGAR", "SANTA BARBARA", "TAMESIS", "TARSO", "TITIRIBI", "URRAO", "VALPARAISO", "VENECIA"],
    "ÁREA METROPOLITANA": ["MEDELLÍN", "BELLO", "ENVIGADO", "ITAGÜÍ", "SABANETA", "LA ESTRELLA", "CALDAS", "COPACABANA", "GIRARDOTA", "BARBOSA"],
}

# ──────────────────────────────────────────────────────────────────────────────
#  FUENTES — Sistema de tres niveles (trust tiers)
#  Tier 3 (TRUST_OFICIAL)   = Gobierno/instituciones → máxima confianza
#  Tier 2 (TRUST_NACIONAL)  = Medios nacionales verificados → alta
#  Tier 1 (TRUST_REGIONAL)  = Medios regionales/locales → media
#  El ranking final prioriza Tier 3 > Tier 2 > Tier 1 antes que el score Tavily.
# ──────────────────────────────────────────────────────────────────────────────
FUENTES_OFICIALES = [
    "policia.gov.co", "invias.gov.co", "mindefensa.gov.co",
    "antioquia.gov.co", "medellin.gov.co", "gobantioquia.gov.co",
    "ideam.gov.co", "ungrd.gov.co", "minenergia.gov.co",
    "fiscalia.gov.co", "presidencia.gov.co", "mintransporte.gov.co",
    "defensoria.gov.co", "procuraduria.gov.co",
]

MEDIOS_NACIONALES = [
    "elcolombiano.com", "eltiempo.com", "semana.com",
    "caracol.com.co", "rcnradio.com", "noticias.caracoltv.com",
    "elespectador.com", "lafm.com.co", "bluradio.com", "wradio.com.co",
    "vanguardia.com", "elheraldo.co", "elmundo.com", "elpais.com.co",
    "publimetro.co", "infobae.com",
]

MEDIOS_REGIONALES = [
    "teleantioquia.co", "minuto30.com", "diarioriente.com",
    "hsbnoticias.com", "extra.com.co", "colombiainforma.info",
    "contagioradio.com", "elmundo.com",
    "h13n.com", "conexionsur.co", "alerta.com.co",
]

# Lista completa para Tavily (include_domains) — unión deduplicada
MEDIOS_COLOMBIANOS = sorted(set(FUENTES_OFICIALES + MEDIOS_NACIONALES + MEDIOS_REGIONALES))

TRUST_OFICIAL  = 3
TRUST_NACIONAL = 2
TRUST_REGIONAL = 1

def tier_de(url: str) -> int:
    """Devuelve el nivel de confianza de una URL según su dominio."""
    if not url:
        return TRUST_REGIONAL
    u = url.lower()
    for d in FUENTES_OFICIALES:
        if d in u:
            return TRUST_OFICIAL
    for d in MEDIOS_NACIONALES:
        if d in u:
            return TRUST_NACIONAL
    return TRUST_REGIONAL

TIER_BADGE = {
    TRUST_OFICIAL:  ("OFICIAL",   "#008751"),
    TRUST_NACIONAL: ("NACIONAL",  "#0ea5e9"),
    TRUST_REGIONAL: ("REGIONAL",  "#64748b"),
}

# Mapping dominio → nombre legible del medio.
# Critical: para que "Fuente: X" en el texto SIEMPRE corresponda a una URL real.
MEDIO_NOMBRE = {
    # Oficiales
    "alertastempranas.defensoria.gov.co": "Defensoría — Alertas Tempranas",
    "defensoria.gov.co":   "Defensoría del Pueblo",
    "policia.gov.co":      "Policía Nacional",
    "invias.gov.co":       "INVÍAS",
    "invias-viajero.vercel.app": "INVÍAS Viajero",
    "mindefensa.gov.co":   "Ministerio de Defensa",
    "antioquia.gov.co":    "Gobernación de Antioquia",
    "gobantioquia.gov.co": "Gobernación de Antioquia",
    "medellin.gov.co":     "Alcaldía de Medellín",
    "ideam.gov.co":        "IDEAM",
    "ungrd.gov.co":        "UNGRD",
    "minenergia.gov.co":   "Min. de Energía",
    "fiscalia.gov.co":     "Fiscalía General",
    "presidencia.gov.co":  "Presidencia",
    "mintransporte.gov.co": "Min. de Transporte",
    "procuraduria.gov.co": "Procuraduría",
    "epm.com.co":          "EPM",
    # Nacionales
    "elcolombiano.com":         "El Colombiano",
    "eltiempo.com":             "El Tiempo",
    "semana.com":               "Semana",
    "noticias.caracoltv.com":   "Noticias Caracol",
    "caracol.com.co":           "Caracol Radio",
    "rcnradio.com":             "RCN Radio",
    "elespectador.com":         "El Espectador",
    "lafm.com.co":              "La FM",
    "bluradio.com":             "Blu Radio",
    "wradio.com.co":            "W Radio",
    "vanguardia.com":           "Vanguardia",
    "elheraldo.co":             "El Heraldo",
    "elmundo.com":              "El Mundo",
    "elpais.com.co":            "El País Cali",
    "publimetro.co":            "Publimetro",
    "infobae.com":              "Infobae",
    # Regionales
    "teleantioquia.co":         "Teleantioquia",
    "minuto30.com":             "Minuto30",
    "diarioriente.com":         "Diariente",
    "h13n.com":                 "H13 Noticias",
    "conexionsur.co":           "Conexión Sur",
    "alerta.com.co":            "Alerta Paisa",
    "hsbnoticias.com":          "HSB Noticias",
    "extra.com.co":             "Extra",
    "colombiainforma.info":     "Colombia Informa",
    "contagioradio.com":        "Contagio Radio",
}

def nombre_medio(url: str) -> str:
    """Devuelve el nombre legible del medio dado un URL. Si no matchea, intenta
    extraer el dominio. Crítico para que el LLM cite fuentes verídicas."""
    if not url:
        return "Fuente desconocida"
    u = url.lower()
    # Match dominios más largos primero (subdominios específicos antes que el genérico)
    for dom in sorted(MEDIO_NOMBRE, key=len, reverse=True):
        if dom in u:
            return MEDIO_NOMBRE[dom]
    m = re.search(r"https?://(?:www\.)?([^/]+)", u)
    if m:
        host = m.group(1)
        return host.split(".")[0].capitalize()
    return "Fuente"

COLORES_EXCEL = {
    "GRAVES":     "FF0000",
    "RELEVANTES": "FFC000",
    "PARCIALES":  "FFFF00",
    "NINGUNA":    "92D050",  # Verde = sin novedad
}

LABELS_CRIT = {
    "GRAVES":     "AFECTACIONES GRAVES",
    "RELEVANTES": "AFECTACIONES RELEVANTES",
    "PARCIALES":  "AFECTACIONES PARCIALES",
    "NINGUNA":    "NO HAY AFECTACIONES",
}

SEMAFORO = {"GRAVES": "🔴", "RELEVANTES": "🟠", "PARCIALES": "🟡", "NINGUNA": "🟢"}

# Reemplazo profesional del semáforo emoji: punto de color CSS + pill tipográfica.
_CRIT_SLUG = {"GRAVES": "graves", "RELEVANTES": "relevantes", "PARCIALES": "parciales", "NINGUNA": "ninguna"}

def dot(crit: str) -> str:
    """Punto de estado coloreado (HTML) según criticidad."""
    return f'<span class="sdot sdot-{_CRIT_SLUG.get(crit, "ninguna")}"></span>'

def pill(crit: str) -> str:
    """Pill tipográfica con la etiqueta de criticidad."""
    slug = _CRIT_SLUG.get(crit, "ninguna")
    return f'<span class="crit-pill pill-{slug}">{LABELS_CRIT.get(crit, "")}</span>'

# Mapeo para fecha en español
DIAS = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO", "DOMINGO"]
MESES = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

def fecha_espanol(dt):
    d = DIAS[dt.weekday()]
    m = MESES[dt.month - 1]
    return f"{d} {dt.day:02d} DE {m}, {dt.year}"

# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS DE RELEVANCIA — núcleo del filtrado anti-basura
# ══════════════════════════════════════════════════════════════════════════════
def _normalize(s: str) -> str:
    """lowercase + sin tildes para comparación robusta."""
    if not s:
        return ""
    return "".join(
        c for c in unicodedata.normalize("NFD", s.lower())
        if unicodedata.category(c) != "Mn"
    )

# Set de tokens válidos para considerar que una noticia es "de Antioquia"
ANTIOQUIA_KEYWORDS = {
    "antioquia", "valle de aburra", "aburra", "uraba", "magdalena medio",
    "bajo cauca", "suroeste antioqueno", "oriente antioqueno",
    "occidente antioqueno", "norte antioqueno", "nordeste antioqueno",
}

# Municipios multi-palabra (frase) y mono-palabra (token exacto)
_MUN_MULTI = set()
_MUN_SINGLE = set()
for _sub, _ms in SUBREGIONES.items():
    _MUN_MULTI.add(_normalize(_sub))
    for _m in _ms:
        nm = _normalize(_m)
        (_MUN_MULTI if " " in nm else _MUN_SINGLE).add(nm)

# Departamentos colombianos donde una noticia NO debería ser asignada a Antioquia
# (si el texto menciona estos y no menciona Antioquia, lo más probable es que sea irrelevante)
OTROS_DEPARTAMENTOS = {
    "cundinamarca", "boyaca", "santander", "norte de santander", "huila",
    "tolima", "meta", "caqueta", "putumayo", "narino", "cauca",
    "valle del cauca", "choco", "cordoba", "sucre", "bolivar", "atlantico",
    "magdalena", "cesar", "guajira", "casanare", "arauca", "vichada",
    "guaviare", "vaupes", "amazonas", "guainia", "risaralda", "caldas", "quindio",
}

# ──────────────────────────────────────────────────────────────────────────────
#  Mapa municipio → subregión (para pre-asignación, evita depender 100% de la IA)
# ──────────────────────────────────────────────────────────────────────────────
_MUN_A_SUB_MULTI: list[tuple[str, str]] = []   # frases (priorizadas: más largas primero)
_MUN_A_SUB_SINGLE: dict[str, str] = {}         # tokens exactos
for _sub, _ms in SUBREGIONES.items():
    _MUN_A_SUB_MULTI.append((_normalize(_sub), _sub))
    for _m in _ms:
        nm = _normalize(_m)
        if " " in nm:
            _MUN_A_SUB_MULTI.append((nm, _sub))
        else:
            _MUN_A_SUB_SINGLE[nm] = _sub
# Más largas primero para evitar matches parciales
_MUN_A_SUB_MULTI.sort(key=lambda x: -len(x[0]))

def detectar_subregion(texto: str) -> str | None:
    """Detecta la subregión de un texto por mención de municipio. None si no determina."""
    n = _normalize(texto)
    if not n:
        return None
    for frase, sub in _MUN_A_SUB_MULTI:
        if frase in n:
            return sub
    tokens = set(re.findall(r"[a-z]+", n))
    for tok in tokens:
        if tok in _MUN_A_SUB_SINGLE:
            return _MUN_A_SUB_SINGLE[tok]
    return None

# ──────────────────────────────────────────────────────────────────────────────
#  Keywords críticas — boost automático de criticidad y de score
#  Basadas en el vocabulario REAL de los informes Comfenalco
# ──────────────────────────────────────────────────────────────────────────────
KEYWORDS_GRAVES = [
    # Orden público — declaraciones formales y hechos extremos
    "toque de queda", "ley seca", "alerta roja", "estado de emergencia",
    "masacre", "atentado", "ataque con explosivos", "ataque armado",
    "ataque sicarial", "sicariato", "carro bomba", "explosion",
    "desplazamiento masivo", "desplazamiento forzado",
    "secuestro", "asesinato", "asesinados", "homicidio multiple",
    "hostigamiento armado", "incursion armada", "paro armado",
    "consejo de seguridad urgente",
    # Vías
    "cierre total", "via cerrada", "carretera cerrada",
    "perdida total de banca", "colapso de puente",
    # Servicios
    "corte total", "vandalizaron",
]
KEYWORDS_RELEVANTES = [
    # Orden público
    "alerta de seguridad", "alerta naranja", "consejo de seguridad",
    "amenazas a campesinos", "amenazas colectivas", "extorsion",
    "amenaza", "atraco", "paro civico", "bloqueo",
    "ajuste de cuentas", "clan del golfo", "el mesa", "eln", "emc",
    "disidencias", "homicidios", "incremento de homicidios",
    "desplazamiento", "familias desplazadas",
    # Vías
    "cierre parcial", "paso a un carril", "paso restringido",
    "cierres intermitentes", "perdida parcial de banca", "deslizamiento",
    "derrumbe", "bloqueo via", "via afectada",
    # Servicios
    "corte de agua", "corte de energia", "interrupcion de energia",
    "interrupcion del servicio", "sin servicio de agua",
    "afectacion de servicios",
]
KEYWORDS_PARCIALES = [
    # Marcadores de seguimiento
    "seguimiento", "continua",
    # Vías programadas
    "obra programada", "mantenimiento", "cierres nocturnos",
    "trabajos en la via", "obras en la via", "obras del metro",
    "perforacion", "pilotes",
    # Alertas blandas
    "alerta amarilla", "alerta preventiva", "creciente del rio",
]

# Keywords NEGATIVAS — temas que NO le interesan a Comfenalco
# Si una noticia matchea muchas de estas y pocas críticas → descarta
KEYWORDS_EXCLUSION = [
    # Deportes y fútbol
    "atletico nacional", "independiente medellin", "envigado fc",
    "liga betplay", "copa libertadores", "futbol", "partido", "gol",
    "selección colombia", "campeon", "torneo",
    # Farándula y entretenimiento
    "farandula", "concierto", "festival", "evento musical", "artista",
    "feria de las flores", "alumbrado",
    # Gastronomía / cultura
    "restaurante", "gastronomia", "receta", "turismo recomendado",
    # Política nacional sin Antioquia
    "petro presidente", "congreso de la republica", "reforma tributaria",
    "reforma pensional", "reforma laboral", "ministro",
    # Economía/bolsa
    "bolsa de valores", "dolar hoy", "criptomoneda", "acciones",
    # Tecnología generica
    "inteligencia artificial", "chatgpt", "iphone", "android",
    # Internacional
    "estados unidos", "europa", "venezuela elecciones", "ucrania",
]

def detectar_criticidad_keywords(texto: str) -> str | None:
    """Devuelve el nivel de criticidad sugerido por keywords, o None si no detecta."""
    n = _normalize(texto)
    if not n:
        return None
    if any(k in n for k in KEYWORDS_GRAVES):
        return "GRAVES"
    if any(k in n for k in KEYWORDS_RELEVANTES):
        return "RELEVANTES"
    if any(k in n for k in KEYWORDS_PARCIALES):
        return "PARCIALES"
    return None

def es_ruido_comfenalco(texto: str) -> bool:
    """
    Filtro anti-ruido: descarta noticias que claramente NO le interesan a Comfenalco
    (deportes, farándula, política nacional sin Antioquia, etc).
    Solo descarta si:
      - matchea ≥2 keywords de exclusión, Y
      - NO matchea ninguna keyword crítica (grave/relevante).
    Así una noticia de "Atlético Nacional" en un atentado SÍ pasaría.
    """
    n = _normalize(texto)
    if not n:
        return False
    hits_exclusion = sum(1 for k in KEYWORDS_EXCLUSION if k in n)
    if hits_exclusion < 3:
        return False
    # Si tiene señales críticas, NO es ruido (puede ser relevante igual)
    if any(k in n for k in KEYWORDS_GRAVES + KEYWORDS_RELEVANTES):
        return False
    return True

def menciona_antioquia(texto: str) -> bool:
    """True si el texto menciona Antioquia o algún municipio antioqueño."""
    n = _normalize(texto)
    if not n:
        return False
    # Frases (multi-palabra)
    for frase in ANTIOQUIA_KEYWORDS | _MUN_MULTI:
        if frase in n:
            return True
    # Tokens exactos (mono-palabra) — usar word boundaries
    tokens = set(re.findall(r"[a-z]+", n))
    if tokens & _MUN_SINGLE:
        return True
    return False

def es_de_otro_departamento(texto: str) -> bool:
    """True si el texto habla claramente de otro departamento colombiano."""
    n = _normalize(texto)
    if not n:
        return False
    for dep in OTROS_DEPARTAMENTOS:
        if dep in n:
            return True
    return False

def extraer_fecha_url(url: str):
    """Saca fecha de URLs tipo /2026/05/22/ o /2026-05-22/."""
    if not url:
        return None
    m = re.search(r"(20\d{2})[/_-](\d{1,2})[/_-](\d{1,2})", url)
    if not m:
        return None
    try:
        y, mo, d = map(int, m.groups())
        return datetime(y, mo, d, 23, 59, 59)
    except (ValueError, TypeError):
        return None

def parsear_fecha_publicacion(fecha_pub: str, url: str):
    """Intenta parsear la fecha de publicación. Devuelve datetime o None."""
    if fecha_pub:
        raw = fecha_pub.strip()
        # 1) ISO (topic="general" suele venir así, p.ej. "2026-05-30")
        try:
            if len(raw) <= 10:
                return datetime.fromisoformat(raw) + timedelta(hours=23, minutes=59, seconds=59)
            return datetime.fromisoformat(raw[:19])
        except (ValueError, TypeError):
            pass
        # 2) RFC 2822 / HTTP-date (topic="news" devuelve "Sun, 30 Aug 2009 00:00:00 GMT")
        try:
            dt = parsedate_to_datetime(raw)
            if dt is not None:
                return dt.replace(tzinfo=None)  # naive, para comparar con datetime.now()
        except (ValueError, TypeError):
            pass
    # Fallback: probar desde el URL
    return extraer_fecha_url(url)


# ══════════════════════════════════════════════════════════════════════════════
#  CSS — diseño minimalista, performante. Senior-grade.
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=Instrument+Serif:ital@0;1&family=JetBrains+Mono:wght@500&display=swap');

:root {
    --brand: #008751;
    --brand-2: #006a40;
    --brand-light: #BBCE00;
    --ink-1: #0b1220;
    --ink-2: #1a2233;
    --ink-3: #475569;
    --ink-4: #64748b;
    --ink-5: #94a3b8;
    --bg: #fafbfc;
    --card: #ffffff;
    --line: #eef0f3;
    --line-2: #e4e7ec;
    --r-lg: 20px;
    --r-md: 14px;
    --r-sm: 10px;
    --sh-1: 0 1px 2px rgba(11, 18, 32, 0.04);
    --sh-2: 0 4px 12px -2px rgba(11, 18, 32, 0.06), 0 2px 4px -1px rgba(11, 18, 32, 0.04);
    --sh-3: 0 12px 32px -8px rgba(11, 18, 32, 0.10), 0 4px 8px -2px rgba(11, 18, 32, 0.05);
    --ease: cubic-bezier(0.22, 1, 0.36, 1);
    --crit-r: #ef4444;
    --crit-o: #f59e0b;
    --crit-y: #eab308;
    --crit-g: #22c55e;
}

/* ── App canvas ──────────────────────────────────────────────────────────── */
.stApp {
    background: var(--bg) !important;
}
.block-container {
    padding-top: 2rem !important;
    padding-bottom: 4rem !important;
    max-width: 1320px !important;
}
#MainMenu, footer, header[data-testid="stHeader"] { display: none !important; }

html, body, [class*="css"], .stMarkdown, p, span, label, li, div {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
    color: var(--ink-2) !important;
    -webkit-font-smoothing: antialiased;
    -moz-osx-font-smoothing: grayscale;
}

h1, h2, h3, h4 {
    font-family: 'Inter', sans-serif !important;
    letter-spacing: -0.02em !important;
    color: var(--ink-1) !important;
    font-weight: 700 !important;
}

@keyframes fadeUp { from{opacity:0;transform:translateY(8px)} to{opacity:1;transform:translateY(0)} }
@keyframes fadeIn { from{opacity:0} to{opacity:1} }
@keyframes softPulse {
    0%,100% { opacity: 1; transform: scale(1); }
    50%     { opacity: 0.55; transform: scale(0.92); }
}

/* ── HERO — minimalista, premium ─────────────────────────────────────────── */
.hero {
    position: relative;
    overflow: hidden;
    border-radius: var(--r-lg);
    padding: 2.4rem 2.6rem;
    margin: 0 0 1.5rem 0;
    background:
        radial-gradient(900px 320px at 100% 0%, rgba(187, 206, 0, 0.22) 0%, transparent 60%),
        linear-gradient(135deg, #03281c 0%, #064a32 45%, #006a44 100%);
    color: white !important;
    box-shadow: 0 20px 40px -16px rgba(3, 40, 28, 0.35);
    animation: fadeIn 0.4s var(--ease);
}
.hero-row {
    position: relative;
    z-index: 2;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 2rem;
}
.hero-left { display: flex; align-items: center; gap: 1.5rem; min-width: 0; flex: 1; }
.hero-logo {
    flex: 0 0 72px;
    width: 72px; height: 72px;
    border-radius: 18px;
    background: #ffffff;
    display: flex; align-items: center; justify-content: center;
    padding: 12px;
    box-shadow: 0 8px 24px -6px rgba(0,0,0,0.3), inset 0 1px 0 rgba(255,255,255,1);
}
.hero-logo img {
    width: 100%; height: 100%;
    object-fit: contain;
    display: block;
}
.hero-text { min-width: 0; }
.hero-tag {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    font-size: 0.7rem;
    font-weight: 600;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    color: rgba(187, 206, 0, 0.95) !important;
    margin: 0 0 0.5rem 0;
}
.hero-tag .dot {
    width: 6px; height: 6px;
    border-radius: 50%;
    background: #BBCE00;
    animation: softPulse 2s ease-in-out infinite;
}
.hero-title {
    font-size: 2.1rem;
    font-weight: 700;
    line-height: 1.05;
    margin: 0;
    color: #ffffff !important;
    letter-spacing: -0.03em;
}
.hero-title em {
    font-family: 'Instrument Serif', serif !important;
    font-style: italic;
    font-weight: 400;
    color: #d8f5e6 !important;
    letter-spacing: -0.01em;
}
.hero-date {
    text-align: right;
    color: rgba(255, 255, 255, 0.65) !important;
    font-size: 0.78rem;
    font-weight: 500;
    letter-spacing: 0.3px;
    text-transform: uppercase;
    white-space: nowrap;
}
.hero-date .day {
    display: block;
    font-family: 'Inter', sans-serif;
    font-size: 1.6rem;
    font-weight: 700;
    color: #ffffff !important;
    letter-spacing: -0.02em;
    margin-bottom: 2px;
    text-transform: none;
}

/* ── Sidebar ─────────────────────────────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: #ffffff !important;
    border-right: 1px solid var(--line);
}
.panel-title {
    font-size: 0.68rem;
    font-weight: 700;
    color: var(--ink-4) !important;
    text-transform: uppercase;
    letter-spacing: 1.8px;
    margin: 0.8rem 0 1rem 0;
}

/* ── Metrics ─────────────────────────────────────────────────────────────── */
div[data-testid="metric-container"] {
    background: var(--card) !important;
    border: 1px solid var(--line) !important;
    border-radius: var(--r-md) !important;
    padding: 1.2rem 1.3rem !important;
    box-shadow: var(--sh-1) !important;
    transition: border-color 0.2s var(--ease), transform 0.2s var(--ease);
}
div[data-testid="metric-container"]:hover {
    border-color: var(--line-2) !important;
    transform: translateY(-2px);
}
div[data-testid="stMetricValue"] > div {
    font-size: 2rem !important;
    font-weight: 700 !important;
    color: var(--ink-1) !important;
    letter-spacing: -0.03em;
    font-variant-numeric: tabular-nums;
}
div[data-testid="stMetricLabel"] > label {
    font-size: 0.7rem !important;
    font-weight: 600 !important;
    color: var(--ink-4) !important;
    text-transform: uppercase;
    letter-spacing: 1.2px;
}

/* ── Botón principal ─────────────────────────────────────────────────────── */
div[data-testid="stButton"] > button {
    background: linear-gradient(180deg, var(--brand) 0%, var(--brand-2) 100%) !important;
    color: white !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 0.9rem 1.6rem !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.88rem !important;
    letter-spacing: 0.3px;
    width: 100% !important;
    box-shadow: 0 6px 16px -6px rgba(0, 135, 81, 0.45),
                inset 0 1px 0 rgba(255, 255, 255, 0.12) !important;
    transition: transform 0.15s var(--ease), box-shadow 0.2s var(--ease) !important;
}
div[data-testid="stButton"] > button:hover {
    transform: translateY(-1px);
    box-shadow: 0 10px 22px -6px rgba(0, 135, 81, 0.5),
                inset 0 1px 0 rgba(255, 255, 255, 0.15) !important;
}
div[data-testid="stButton"] > button:active { transform: translateY(0); }

div[data-testid="stDownloadButton"] > button {
    background: var(--ink-1) !important;
    color: white !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 0.85rem 1.6rem !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    letter-spacing: 0.3px;
    box-shadow: 0 4px 12px -3px rgba(11, 18, 32, 0.25) !important;
    transition: transform 0.15s var(--ease) !important;
}
div[data-testid="stDownloadButton"] > button:hover { transform: translateY(-1px); }

/* ── Expanders ───────────────────────────────────────────────────────────── */
div[data-testid="stExpander"] {
    background: var(--card) !important;
    border: 1px solid var(--line) !important;
    border-radius: var(--r-md) !important;
    margin-bottom: 0.6rem !important;
    box-shadow: var(--sh-1) !important;
    overflow: hidden;
    transition: border-color 0.2s var(--ease);
}
div[data-testid="stExpander"]:hover { border-color: var(--line-2) !important; }
div[data-testid="stExpander"] summary {
    padding: 0.9rem 1.1rem !important;
    font-weight: 600 !important;
    font-size: 0.9rem;
}
.crit-graves     div[data-testid="stExpander"] { border-left: 3px solid var(--crit-r) !important; }
.crit-relevantes div[data-testid="stExpander"] { border-left: 3px solid var(--crit-o) !important; }
.crit-parciales  div[data-testid="stExpander"] { border-left: 3px solid var(--crit-y) !important; }
.crit-ninguna    div[data-testid="stExpander"] { border-left: 3px solid var(--crit-g) !important; }

/* ── Status dots — reemplazo profesional del semáforo emoji ───────────────── */
.sdot {
    display: inline-block;
    width: 9px; height: 9px;
    border-radius: 50%;
    margin-right: 7px;
    vertical-align: middle;
    position: relative;
    top: -1px;
}
.sdot-graves     { background: var(--crit-r); box-shadow: 0 0 0 3px rgba(239,68,68,0.15); }
.sdot-relevantes { background: var(--crit-o); box-shadow: 0 0 0 3px rgba(245,158,11,0.15); }
.sdot-parciales  { background: var(--crit-y); box-shadow: 0 0 0 3px rgba(234,179,8,0.15); }
.sdot-ninguna    { background: var(--crit-g); box-shadow: 0 0 0 3px rgba(34,197,94,0.15); }
/* Pill de criticidad para encabezados de subregión */
.crit-pill {
    display: inline-block;
    font-size: 0.62rem;
    font-weight: 700;
    letter-spacing: 0.8px;
    text-transform: uppercase;
    padding: 3px 9px;
    border-radius: 999px;
    vertical-align: middle;
}
.pill-graves     { background: rgba(239,68,68,0.12); color: #b91c1c !important; }
.pill-relevantes { background: rgba(245,158,11,0.14); color: #b45309 !important; }
.pill-parciales  { background: rgba(234,179,8,0.16); color: #a16207 !important; }
.pill-ninguna    { background: rgba(34,197,94,0.12); color: #15803d !important; }

/* Acento superior en metric cards — toque institucional */
div[data-testid="metric-container"] {
    position: relative;
    overflow: hidden;
}
div[data-testid="metric-container"]::before {
    content: "";
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, var(--brand), var(--brand-light));
    opacity: 0.85;
}

/* ── Stat cards — resumen de criticidad (reemplaza st.metric con emoji) ──── */
.stat-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 0.75rem;
    margin: 0.4rem 0 0.2rem 0;
}
.stat-card {
    background: var(--card);
    border: 1px solid var(--line);
    border-top: 3px solid var(--sc, var(--brand));
    border-radius: var(--r-md);
    padding: 1.05rem 1.2rem;
    box-shadow: var(--sh-1);
    transition: transform 0.2s var(--ease), border-color 0.2s var(--ease);
}
.stat-card:hover { transform: translateY(-2px); border-color: var(--line-2); }
.stat-num {
    font-size: 2.1rem;
    font-weight: 700;
    color: var(--ink-1) !important;
    letter-spacing: -0.03em;
    font-variant-numeric: tabular-nums;
    line-height: 1;
}
.stat-lab {
    margin-top: 0.45rem;
    font-size: 0.7rem;
    font-weight: 600;
    color: var(--ink-4) !important;
    text-transform: uppercase;
    letter-spacing: 1.2px;
    display: flex; align-items: center;
}

/* ── Source tag ──────────────────────────────────────────────────────────── */
.source-tag {
    background: #f5f7fa;
    color: var(--brand) !important;
    padding: 4px 9px;
    border-radius: 6px;
    font-size: 0.68rem;
    font-weight: 600;
    text-decoration: none !important;
    border: 1px solid var(--line);
    transition: background 0.15s var(--ease), color 0.15s var(--ease);
    display: inline-block;
    margin: 2px 4px 2px 0;
}
.source-tag:hover {
    background: var(--brand);
    color: white !important;
    border-color: var(--brand);
}

/* ── News card ───────────────────────────────────────────────────────────── */
.news-card {
    background: var(--card);
    border: 1px solid var(--line);
    border-left: 3px solid var(--brand);
    border-radius: 10px;
    padding: 0.85rem 1rem;
    margin-bottom: 0.55rem;
    transition: border-color 0.15s var(--ease);
}
.news-card:hover { border-color: var(--line-2); border-left-color: var(--brand-light); }
.news-meta {
    font-size: 0.7rem;
    color: var(--ink-4) !important;
    font-weight: 500;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 0.35rem;
}
.news-title {
    font-size: 0.92rem;
    font-weight: 600;
    color: var(--ink-1) !important;
    margin: 0 0 0.35rem 0;
    line-height: 1.4;
    letter-spacing: -0.005em;
}
.news-snippet {
    font-size: 0.8rem;
    color: var(--ink-3) !important;
    line-height: 1.55;
    margin: 0 0 0.5rem 0;
}

/* ── Checkboxes ──────────────────────────────────────────────────────────── */
div[data-testid="stCheckbox"] {
    border-radius: 8px;
    padding: 2px 10px !important;
    margin-left: -10px !important;
    width: calc(100% + 20px) !important;
    transition: background 0.15s var(--ease);
}
div[data-testid="stCheckbox"]:hover { background: #f5f7fa !important; cursor: pointer; }
div[data-testid="stCheckbox"] label p {
    font-size: 0.82rem !important;
    color: var(--ink-3) !important;
    font-weight: 500;
}

/* ── Progress bar — sin shimmer infinito (perf) ──────────────────────────── */
.stProgress > div > div {
    background: var(--line) !important;
    border-radius: 99px !important;
    height: 6px !important;
}
.stProgress > div > div > div {
    background: linear-gradient(90deg, var(--brand), var(--brand-light)) !important;
    height: 6px !important;
    border-radius: 99px !important;
    transition: width 0.3s var(--ease);
}

/* ── Placeholder — minimal ──────────────────────────────────────────────── */
.placeholder {
    background: var(--card);
    border: 1px solid var(--line);
    border-radius: var(--r-lg);
    padding: 4.5rem 2rem;
    text-align: center;
    animation: fadeIn 0.3s var(--ease);
}
.placeholder-glyph {
    width: 48px; height: 48px;
    margin: 0 auto 1.4rem;
    border-radius: 14px;
    background: linear-gradient(135deg, rgba(0,135,81,0.08), rgba(187,206,0,0.10));
    display: flex; align-items: center; justify-content: center;
    color: var(--brand);
}
.placeholder-glyph svg { width: 22px; height: 22px; }
.placeholder h2 {
    font-size: 1.3rem;
    font-weight: 700;
    color: var(--ink-1) !important;
    margin: 0 0 0.4rem 0;
    letter-spacing: -0.02em;
}
.placeholder p {
    color: var(--ink-4) !important;
    font-size: 0.9rem;
    margin: 0 auto;
    max-width: 380px;
    line-height: 1.55;
}

/* ── Section heading ─────────────────────────────────────────────────────── */
.section-h {
    font-size: 0.78rem;
    font-weight: 700;
    color: var(--ink-4) !important;
    margin: 2rem 0 0.9rem 0;
    text-transform: uppercase;
    letter-spacing: 1.6px;
    display: flex;
    align-items: center;
    gap: 10px;
}
.section-h::before {
    content: "";
    width: 16px; height: 2px;
    background: var(--brand);
    border-radius: 2px;
}

/* ── Alerts ──────────────────────────────────────────────────────────────── */
div[data-testid="stAlert"] {
    border-radius: 10px !important;
    border: 1px solid var(--line) !important;
    padding: 0.7rem 0.9rem !important;
    font-size: 0.84rem !important;
}

/* ── Tabs ───────────────────────────────────────────────────────────────── */
div[data-baseweb="tab-list"] {
    gap: 2px !important;
    border-bottom: 1px solid var(--line) !important;
}
button[data-baseweb="tab"] {
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.84rem !important;
    color: var(--ink-4) !important;
    padding: 9px 14px !important;
}
button[data-baseweb="tab"][aria-selected="true"] {
    color: var(--brand) !important;
    box-shadow: inset 0 -2px 0 var(--brand) !important;
}

/* ── Inputs ──────────────────────────────────────────────────────────────── */
div[data-testid="stDateInput"] input,
div[data-testid="stTextInput"] input {
    border-radius: 10px !important;
    border: 1px solid var(--line-2) !important;
    background: white !important;
    padding: 9px 12px !important;
    font-size: 0.88rem !important;
    transition: border-color 0.15s var(--ease), box-shadow 0.15s var(--ease);
}
div[data-testid="stDateInput"] input:focus,
div[data-testid="stTextInput"] input:focus {
    border-color: var(--brand) !important;
    box-shadow: 0 0 0 3px rgba(0, 135, 81, 0.12) !important;
}

/* ── Footer ──────────────────────────────────────────────────────────────── */
.brand-foot {
    margin-top: 3.5rem;
    padding: 1.2rem 0 0.5rem 0;
    border-top: 1px solid var(--line);
    text-align: center;
    font-size: 0.72rem;
    color: var(--ink-5) !important;
    letter-spacing: 0.3px;
}
.brand-foot b { color: var(--ink-3) !important; font-weight: 600; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  BÚSQUEDA — Tavily con filtrado estricto
# ══════════════════════════════════════════════════════════════════════════════
# Estadísticas de descarte por query (para diagnóstico)
_SEARCH_STATS: dict[str, dict] = {}

def buscar_noticias(query: str, tavily_key: str, max_resultados: int = 15, cat_label: str = "?") -> tuple[list[dict], dict]:
    """
    Devuelve (noticias_filtradas, stats) — sub-query atómica.
    Filtros aplicados:
      • Fecha real ≤ 7 días (parsed de published_date o de URL; si no se puede, marca aprox).
      • Mención de Antioquia o algún municipio antioqueño.
      • Anti-ruido Comfenalco (≥3 keywords de exclusión y sin señal crítica).
      • Dedupe por título normalizado.
      • Ranking por score compuesto numérico.
    El caller agrega múltiples sub-queries en _SEARCH_STATS[categoría].
    """
    stats = {"raw": 0, "no_basics": 0, "fecha_aprox": 0, "too_old": 0, "no_geo": 0, "ruido": 0, "dup": 0, "kept": 0}
    try:
        client = TavilyClient(api_key=tavily_key)
        # topic="news" + days=7 → el índice de noticias de Tavily devuelve
        # published_date de forma mucho más confiable que topic="general", lo que
        # reduce drásticamente las noticias marcadas "fecha-aprox" (penalizadas) y
        # mejora la frescura real del ranking. Mantiene include_domains (trust tiers).
        resp = client.search(
            query=query,
            search_depth="advanced",
            max_results=max_resultados * 3,
            topic="news",
            days=7,
            include_domains=MEDIOS_COLOMBIANOS,
        )

        # Corte alineado con time_range="week" (7 días); el score compuesto se encarga
        # de empujar arriba lo más reciente y oficial.
        corte = datetime.now() - timedelta(days=7)
        candidatos = []
        raw_results = resp.get("results", [])
        stats["raw"] = len(raw_results)

        for r in raw_results:
            titulo    = (r.get("title") or "").strip()
            resumen   = (r.get("content") or "").strip()[:600]
            url       = (r.get("url") or "").strip()
            fecha_pub = r.get("published_date", "")
            score     = float(r.get("score", 0.0) or 0.0)

            # Higiene mínima
            if not titulo or len(resumen) < 60 or not url:
                stats["no_basics"] += 1
                continue

            # Fecha — Tavily ya filtró por time_range="week"; si no hay fecha exacta
            # aceptamos con marca "aprox" y penalización al score.
            dt = parsear_fecha_publicacion(fecha_pub, url)
            fecha_aprox = False
            if dt is None:
                # Asumir mid-week (3.5 días atrás) para ranking; marcar aprox
                dt = datetime.now() - timedelta(days=3, hours=12)
                fecha_aprox = True
                stats["fecha_aprox"] += 1
            elif dt < corte:
                stats["too_old"] += 1
                continue

            texto_completo = f"{titulo} {resumen}"

            # Relevancia geográfica
            if not menciona_antioquia(texto_completo):
                stats["no_geo"] += 1
                continue

            # Filtro anti-ruido (farándula/deportes/política nacional sin valor)
            if es_ruido_comfenalco(texto_completo):
                stats["ruido"] += 1
                continue

            tier = tier_de(url)
            sub_detectada   = detectar_subregion(texto_completo)
            criticidad_hint = detectar_criticidad_keywords(texto_completo)

            # Score compuesto numérico (más alto = más prioritario):
            #   tier_weight (max 300) + score_tavily (max 100) + recency_boost (max 100)
            #   + criticidad_boost (GRAVES=80, RELEVANTES=40, PARCIALES=10)
            #   - penalty fecha_aprox (-40 para que confirmadas salgan arriba)
            horas_atras = (datetime.now() - dt).total_seconds() / 3600
            recency_boost = max(0, 100 - horas_atras)
            crit_boost = {"GRAVES": 80, "RELEVANTES": 40, "PARCIALES": 10}.get(criticidad_hint or "", 0)
            penalty_aprox = -40 if fecha_aprox else 0
            score_compuesto = (tier * 100) + (score * 100) + recency_boost + crit_boost + penalty_aprox

            candidatos.append({
                "titulo":          titulo,
                "resumen":         resumen,
                "url":             url,
                "fecha":           "~ esta semana" if fecha_aprox else dt.strftime("%d/%m/%Y"),
                "fecha_dt":        dt,
                "fecha_aprox":     fecha_aprox,
                "score":           score,
                "score_compuesto": score_compuesto,
                "tier":            tier,
                "sub_detectada":   sub_detectada,
                "criticidad_hint": criticidad_hint,
            })

        # Dedupe por título normalizado — al colisionar conserva el de mayor score compuesto
        vistos: dict[str, dict] = {}
        for c in candidatos:
            key = _normalize(c["titulo"])[:80]
            if key not in vistos or c["score_compuesto"] > vistos[key]["score_compuesto"]:
                vistos[key] = c
            else:
                stats["dup"] += 1
        unicos = list(vistos.values())

        # Ranking por score compuesto (numérico, no lexicográfico)
        unicos.sort(key=lambda x: x["score_compuesto"], reverse=True)
        kept = unicos[:max_resultados]
        stats["kept"] = len(kept)
        return kept, stats

    except Exception as e:
        st.warning(f"⚠️ Error Tavily ({cat_label}): {e}")
        return [], stats


def formatear(resultados: list[dict], offset: int = 0) -> tuple[str, list[dict]]:
    """
    Devuelve (texto_para_prompt, lista_indexada).
    Cada noticia recibe un ID [N{n}] + hints de pre-clasificación para que la IA
    no tenga que adivinar la subregión cuando ya el sistema la detectó.
    """
    if not resultados:
        return "(Sin noticias recientes en medios colombianos)", []
    lineas = []
    indexadas = []
    for i, r in enumerate(resultados, start=1 + offset):
        sub_hint  = f" · 📍 SUBREGIÓN_DETECTADA={r['sub_detectada']}" if r.get("sub_detectada") else ""
        crit_hint = f" · 🎯 CRITICIDAD_SUGERIDA={r['criticidad_hint']}" if r.get("criticidad_hint") else ""
        medio = nombre_medio(r["url"])
        lineas.append(
            f"[N{i}] ({r['fecha']}){sub_hint}{crit_hint}\n"
            f"     {r['titulo']}\n"
            f"     {r['resumen']}\n"
            f"     📰 Medio: {medio}\n"
            f"     URL: {r['url']}"
        )
        indexadas.append({**r, "id": i, "medio": medio})
    return "\n\n".join(lineas), indexadas


def _limpiar_texto_oficial(raw: str) -> str:
    """Limpia HTML/menús del raw_content y conserva solo líneas relevantes."""
    if not raw:
        return ""
    # Sacar líneas muy cortas o típicas de navegación
    lineas = [l.strip() for l in raw.splitlines() if l.strip()]
    keywords = ["via", "vía", "carretera", "cierre", "derrumbe", "bloqueo",
                "tramo", "kilómetro", "km ", "paso", "habilitad", "transitabl",
                "movilidad", "antioquia", "medellin", "medellín", "uraba", "urabá"]
    relevantes = []
    for l in lineas:
        nl = _normalize(l)
        if len(l) < 25 or len(l) > 400:
            continue
        if any(k in nl for k in keywords):
            relevantes.append(l)
    return "\n".join(relevantes[:30])


# ──────────────────────────────────────────────────────────────────────────────
#  CIERRES VIALES ACTIVOS — modelo de "validez vigente", no de "publicación reciente"
#  Un cierre dado de alta el 02/04/2026 con fin el 30/07/2026 sigue siendo válido
#  HOY si la fecha actual cae dentro del rango (o si está marcado como indefinido).
#  Estos datos NO se filtran por la ventana de 72h.
# ──────────────────────────────────────────────────────────────────────────────
_FECHA_DDMMYYYY = re.compile(r"\b(\d{1,2})/(\d{1,2})/(20\d{2})\b")

def _parse_ddmmyyyy(d: str, m: str, y: str):
    try:
        return datetime(int(y), int(m), int(d), 23, 59, 59)
    except (ValueError, TypeError):
        return None

def extraer_avisos_vigentes(raw_content: str, fuente_url: str, tipo_evento: str = "OFICIAL") -> list[dict]:
    """
    Extrae avisos 'vigentes hoy' del raw_content de fuentes oficiales.
    Modelo bitemporal: lo que importa es la ventana 'Inicio → Fin' (valid time),
    no la fecha de publicación (transaction time).

    Aplica a:
      • Cierres viales (Policía/INVÍAS)
      • Cortes programados de servicios (EPM, Aguas, EMVarias)
      • Alertas oficiales con ventana (IDEAM, UNGRD)

    Estrategia: ventanas de ~400 chars alrededor de cada mención a 'antioquia'.
    Vigente si la fecha máxima en la ventana ≥ hoy, o si el texto dice 'indefinido'.
    """
    if not raw_content:
        return []

    hoy = datetime.now()
    activos = []
    vistos = set()

    for m in re.finditer(r"antioquia", raw_content, re.IGNORECASE):
        ini = max(0, m.start() - 400)
        fin = min(len(raw_content), m.end() + 400)
        bloque = raw_content[ini:fin]

        firma = re.sub(r"\s+", "", bloque)[:200]
        if firma in vistos:
            continue
        vistos.add(firma)

        fechas = []
        for fm in _FECHA_DDMMYYYY.finditer(bloque):
            dt = _parse_ddmmyyyy(fm.group(1), fm.group(2), fm.group(3))
            if dt:
                fechas.append(dt)

        indefinido = "indefinido" in _normalize(bloque)

        vigente = False
        fecha_fin = None
        if fechas:
            fecha_fin = max(fechas)
            vigente = fecha_fin >= hoy
        if indefinido and fechas:
            vigente = True

        if not vigente:
            continue

        limpio = re.sub(r"\s+", " ", bloque).strip()
        if len(limpio) > 600:
            limpio = limpio[:600] + "…"

        activos.append({
            "tipo_evento":      tipo_evento,
            "texto":            limpio,
            "fecha_fin":        fecha_fin.strftime("%d/%m/%Y") if fecha_fin else "indefinido",
            "indefinido":       indefinido,
            "fuente":           fuente_url,
            "sub_detectada":    detectar_subregion(limpio),
            "criticidad_hint":  detectar_criticidad_keywords(limpio),
        })

    return activos


# Wrappers semánticos
def extraer_cierres_viales_activos(raw_content: str, fuente_url: str) -> list[dict]:
    return extraer_avisos_vigentes(raw_content, fuente_url, tipo_evento="CIERRE_VIAL")

def extraer_avisos_servicios_activos(raw_content: str, fuente_url: str) -> list[dict]:
    return extraer_avisos_vigentes(raw_content, fuente_url, tipo_evento="AVISO_SERVICIO")


def formatear_avisos_vigentes(avisos: list[dict], etiqueta: str, prefijo_id: str = "AV") -> str:
    """Convierte avisos vigentes a bloque de texto para el prompt (con hints de pre-clasificación)."""
    if not avisos:
        return f"(Sin {etiqueta.lower()} activos en fuentes oficiales hoy)"
    lineas = []
    for i, c in enumerate(avisos, start=1):
        marca = "INDEFINIDO" if c["indefinido"] else f"VIGENTE HASTA {c['fecha_fin']}"
        sub_hint  = f" · 📍 SUBREGIÓN_DETECTADA={c['sub_detectada']}" if c.get("sub_detectada") else ""
        crit_hint = f" · 🎯 CRITICIDAD_SUGERIDA={c['criticidad_hint']}" if c.get("criticidad_hint") else ""
        medio = nombre_medio(c["fuente"])
        lineas.append(
            f"[{prefijo_id}{i}] ⚠ {marca}{sub_hint}{crit_hint}\n"
            f"     {c['texto']}\n"
            f"     📰 Medio oficial: {medio}\n"
            f"     URL: {c['fuente']}"
        )
    return "\n\n".join(lineas)

# Alias para compatibilidad
formatear_cierres_activos = lambda c: formatear_avisos_vigentes(c, "cierres viales", "CV")


def obtener_avisos_servicios_oficiales(tavily_key: str) -> tuple[str, list[dict], list[str]]:
    """
    Avisos oficiales de servicios públicos vigentes hoy (cortes programados, alertas).
    Fuentes: EPM, UNGRD, IDEAM, Gobernación de Antioquia, Alcaldía de Medellín.
    Devuelve (texto_libre_limpio, avisos_estructurados, urls_oficiales_usadas).
    """
    urls_oficiales = [
        "https://www.epm.com.co/site/clientes/clientes-y-usuarios",
        "https://www.ungrd.gov.co/sala-de-prensa/comunicados",
        "https://www.ideam.gov.co/web/tiempo-y-clima/alertas-tempranas",
        "https://alertastempranas.defensoria.gov.co/",
        "https://www.antioquia.gov.co/index.php/component/k2/itemlist/category/220-noticias",
        "https://www.medellin.gov.co",
    ]
    try:
        client = TavilyClient(api_key=tavily_key)
        result = client.extract(urls=urls_oficiales)
        bloques = []
        avisos = []
        urls_ok = []
        for r in result.get("results", []):
            url = r.get("url", "")
            raw = r.get("raw_content", "")
            contenido = _limpiar_texto_oficial(raw)
            avisos_aqui = extraer_avisos_servicios_activos(raw, url)
            if contenido or avisos_aqui:
                if contenido:
                    bloques.append(f"=== Fuente oficial: {url} ===\n{contenido}")
                avisos.extend(avisos_aqui)
                urls_ok.append(url)
        texto = "\n\n".join(bloques) if bloques else "(No se pudo acceder a fuentes oficiales de servicios)"
        return texto, avisos, urls_ok
    except Exception as e:
        return f"(Error accediendo fuentes oficiales de servicios: {e})", [], []


def obtener_estado_vias_oficial(tavily_key: str) -> tuple[str, list[dict], list[str]]:
    """
    Extrae estado de vías oficial.
    Devuelve (texto_libre_limpio, cierres_activos_estructurados, urls_oficiales_usadas).
    """
    urls_oficiales = [
        "https://invias-viajero.vercel.app",
        "https://www.policia.gov.co/estado-de-las-vias",
        "https://invias.gov.co/index.php/informacion-institucional/estado-vias",
        "https://www.antioquia.gov.co/index.php/component/k2/itemlist/category/220-noticias",
        "https://www.ideam.gov.co/web/tiempo-y-clima/alertas-tempranas",
    ]
    try:
        client = TavilyClient(api_key=tavily_key)
        result = client.extract(urls=urls_oficiales)
        bloques = []
        cierres_activos = []
        urls_ok = []
        for r in result.get("results", []):
            url = r.get("url", "")
            raw = r.get("raw_content", "")
            # Texto libre limpio (para contexto general)
            contenido = _limpiar_texto_oficial(raw)
            # Cierres estructurados vigentes (modelo de validez, no de publicación)
            cierres = extraer_cierres_viales_activos(raw, url)
            if contenido or cierres:
                if contenido:
                    bloques.append(f"=== Fuente oficial: {url} ===\n{contenido}")
                cierres_activos.extend(cierres)
                urls_ok.append(url)
        texto = "\n\n".join(bloques) if bloques else "(No se pudo acceder a fuentes oficiales)"
        return texto, cierres_activos, urls_ok
    except Exception as e:
        return f"(Error accediendo fuentes oficiales: {e})", [], []


def obtener_noticias(tavily_key: str) -> dict:
    """
    Devuelve un dict con todo el contexto para el prompt y los maps de URLs por categoría:
    {
      "texto_seg": str, "texto_vias": str, "texto_svc": str,
      "indexadas": list[dict],          # todas las noticias con ID
      "url_to_cat": dict[str, str],     # url -> 'seg' | 'vias' | 'svc'
    }
    """
    # Queries Comfenalco: sub-queries cortas y naturales por categoría
    # → Tavily prefiere queries cortas; usamos 3-4 ángulos por categoría para
    #   cubrir distintos sub-temas y luego deduplicamos por URL.
    queries = [
        ("seg", [
            # Ángulo 1: orden público clásico — subregiones de mayor conflicto
            "Antioquia Bajo Cauca Norte Nordeste orden público ataque armado desplazamiento grupo armado",
            # Ángulo 2: sicariato y bandas criminales — Urabá y Valle de Aburrá
            "Antioquia Urabá Medellín ataque sicarial homicidio Clan del Golfo consejo de seguridad",
            # Ángulo 3: declaraciones y crisis — Norte/Occidente/Suroeste
            "Antioquia Ituango Briceño Cañasgordas alerta roja toque de queda masacre desplazamiento familias",
            # Ángulo 4: patrones recurrentes de violencia — Oriente y Nordeste
            "Antioquia Oriente Nordeste ola homicidios asesinatos enfrentamientos bandas extorsión El Mesa",
        ]),
        ("vias", [
            # Ángulo 1: cierres y eventos naturales
            "Antioquia vías cierre carretera derrumbe deslizamiento bloqueo Urabá Bajo Cauca",
            # Ángulo 2: obras, concesiones y rutas
            "Antioquia carretera paso restringido Ruta Nacional INVÍAS concesión obras Túnel La Llorona",
            # Ángulo 3: alertas IDEAM y movilidad
            "Antioquia movilidad alerta IDEAM lluvia río Cauca creciente puente afectación vía",
        ]),
        ("svc", [
            # Ángulo 1: servicios públicos EPM
            "Antioquia corte agua energía EPM servicio público paro cívico",
            # Ángulo 2: cortes programados y mantenimiento
            "Antioquia interrupción servicio mantenimiento programado EPM Aguas",
            # Ángulo 3: alertas oficiales con impacto comunitario
            "Antioquia alerta temprana Defensoría UNGRD comunidad afectación",
        ]),
        ("sed", [
            # Sedes, hoteles y parques de Comfenalco Antioquia (col 4)
            # Ángulo 1: instalaciones Comfenalco directas
            "Comfenalco Antioquia sede parque hotel servicio afectación cierre",
            # Ángulo 2: parques recreativos y turismo regional
            "Antioquia parque recreativo turismo cierre clausura emergencia",
            # Ángulo 3: hoteles y alojamiento turístico
            "Antioquia hotel alojamiento turismo afectación servicio cierre",
        ]),
    ]

    bloques_texto = {}
    todas_indexadas = []
    url_to_cat = {}
    offset = 0
    _SEARCH_STATS.clear()

    primera = True
    for cat, sub_queries in queries:
        # Acumulador por categoría
        cat_results: dict[str, dict] = {}  # url → noticia (dedupe)
        cat_stats = {"raw": 0, "no_basics": 0, "fecha_aprox": 0, "too_old": 0,
                     "no_geo": 0, "ruido": 0, "dup": 0, "kept": 0, "sub_queries": 0}

        for sq in sub_queries:
            if not primera:
                time.sleep(0.4)  # respetar rate-limit Tavily
            primera = False
            crudos, sub_stats = buscar_noticias(
                sq, tavily_key,
                max_resultados=12,
                cat_label=f"{cat}/sub",
            )
            cat_stats["sub_queries"] += 1
            for k, v in sub_stats.items():
                if k in cat_stats:
                    cat_stats[k] += v
            # Dedupe por URL preservando el de mayor score compuesto
            for c in crudos:
                u = c["url"]
                if u not in cat_results or c["score_compuesto"] > cat_results[u]["score_compuesto"]:
                    cat_results[u] = c

        # Re-ranking final por score compuesto, top 15
        unicos = sorted(cat_results.values(),
                        key=lambda x: x["score_compuesto"], reverse=True)[:15]
        cat_stats["kept"] = len(unicos)
        _SEARCH_STATS[cat] = cat_stats

        texto, indexadas = formatear(unicos, offset=offset)
        bloques_texto[cat] = texto
        todas_indexadas.extend(indexadas)
        for it in indexadas:
            url_to_cat[it["url"]] = cat
        offset += len(indexadas)

    # ── Enriquecer vías con estado oficial + cierres vigentes (modelo bitemporal)
    estado_oficial, cierres_activos, urls_oficiales = obtener_estado_vias_oficial(tavily_key)
    cierres_txt = formatear_avisos_vigentes(cierres_activos, "cierres viales", "CV")

    bloques_texto["vias"] = (
        "━━━ CIERRES VIALES OFICIALES VIGENTES HOY ━━━\n"
        "(Datos estructurados de Policía/INVÍAS. Su validez NO depende de la fecha de\n"
        "publicación sino de la ventana 'Inicio Afectación → Fin Afectación' o de\n"
        "estar marcados como 'Cierre indefinido'. Inclúyelos SIEMPRE en la celda de\n"
        "vías de la subregión correspondiente, AUN si fueron registrados hace meses.)\n\n"
        f"{cierres_txt}\n\n"
        "━━━ CONTEXTO OFICIAL ADICIONAL (texto libre INVÍAS/Policía) ━━━\n"
        f"{estado_oficial}\n\n"
        "━━━ NOTICIAS DE MEDIOS (últimas 72h) ━━━\n"
        f"{bloques_texto['vias']}"
    )
    for u in urls_oficiales:
        url_to_cat.setdefault(u, "vias")

    # ── Enriquecer servicios con avisos oficiales vigentes (modelo bitemporal)
    estado_svc, avisos_servicios, urls_svc_oficiales = obtener_avisos_servicios_oficiales(tavily_key)
    avisos_svc_txt = formatear_avisos_vigentes(avisos_servicios, "avisos de servicios", "AS")

    bloques_texto["svc"] = (
        "━━━ AVISOS OFICIALES DE SERVICIOS VIGENTES HOY ━━━\n"
        "(Cortes programados de EPM/Aguas/EMVarias, alertas IDEAM/UNGRD, comunicados\n"
        "de la Gobernación/Alcaldía. Misma regla que los cierres viales: la validez\n"
        "depende de la ventana 'desde → hasta', no de cuándo fue publicado el aviso.\n"
        "Inclúyelos SIEMPRE en la celda de Comercio/Servicios y, si afectan agua o\n"
        "energía, también en la de Sedes/Hoteles/Parques de la subregión correspondiente.)\n\n"
        f"{avisos_svc_txt}\n\n"
        "━━━ CONTEXTO OFICIAL ADICIONAL (texto libre EPM/IDEAM/Gob) ━━━\n"
        f"{estado_svc}\n\n"
        "━━━ NOTICIAS DE MEDIOS (últimas 72h) ━━━\n"
        f"{bloques_texto['svc']}"
    )
    for u in urls_svc_oficiales:
        url_to_cat.setdefault(u, "svc")

    return {
        "texto_seg":   bloques_texto["seg"],
        "texto_vias":  bloques_texto["vias"],
        "texto_svc":   bloques_texto["svc"],
        "texto_sed":   bloques_texto.get("sed", "(Sin noticias específicas de sedes/hoteles/parques)"),
        "indexadas":   todas_indexadas,
        "url_to_cat":  url_to_cat,
        "urls_oficiales_vias": urls_oficiales,
        "urls_oficiales_svc":  urls_svc_oficiales,
        "cierres_activos":     cierres_activos,
        "avisos_servicios":    avisos_servicios,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  IA: Groq Llama 3.3 70B
# ══════════════════════════════════════════════════════════════════════════════
def construir_prompt(seg: str, vias: str, svc: str, sed: str, fecha: str) -> str:
    lista = "\n".join(f"  • {s}: {', '.join(m[:5])}" for s, m in SUBREGIONES.items())
    return f"""Eres el analista de seguridad y movilidad de Comfenalco Antioquia.
Fecha del informe: {fecha}.

SUBREGIONES Y MUNICIPIOS DE ANTIOQUIA:
{lista}

═══════════════════════════════════════════════════════════════════════════════
REGLAS CRÍTICAS (NO LAS VIOLES)
═══════════════════════════════════════════════════════════════════════════════
A. Solo usa noticias que mencionen explícitamente Antioquia o un municipio
   antioqueño. Si una noticia habla de otra región (Bogotá, Cali, Cundinamarca,
   Valle, etc.) y NO menciona Antioquia ni un municipio antioqueño → IGNÓRALA.
B. NO inventes datos. Si una subregión no tiene noticias para una categoría,
   escribe EXACTAMENTE: SIN NOVEDAD
C. No mezcles información de distintas noticias en una misma frase si no son
   del mismo hecho. Mejor sepáralas con doble salto de línea.
D. Para cada celda con contenido, devuelve las URLs reales usadas en el campo
   "fuentes_*" correspondiente (copia textual de las URLs presentes en el contexto).
E. DISTINCIÓN CRÍTICA — TIPOS DE DATO:
   • NOTICIAS (medios) → tienen fecha de publicación; ya fueron filtradas a 72h.
   • AVISOS OFICIALES VIGENTES (cierres viales, cortes de servicios programados,
     alertas IDEAM/UNGRD) → tienen ventana de validez (Inicio → Fin) o marca de
     "indefinido". Un aviso registrado el 02/04/2026 con fin el 30/07/2026 sigue
     siendo VÁLIDO HOY y DEBE aparecer en el informe.
   Nunca descartes un aviso vigente por su fecha de registro.

   Criticidad típica de avisos:
     • Cierre total de vía / corte total de servicio / alerta roja → GRAVES
     • Cierre parcial / paso a un carril / corte prolongado / alerta naranja → RELEVANTES
     • Obras programadas / paso restringido / corte corto con horario / alerta amarilla → PARCIALES

   Reglas de propagación entre celdas:
     • Un AVISO_SERVICIO de corte de agua o energía debe aparecer en
       Comercio/Servicios Y TAMBIÉN en Sedes/Hoteles/Parques de la subregión.
     • Un CIERRE_VIAL grave que aísle un municipio puede elevar también la celda
       de Comercio/Servicios a RELEVANTES (impacto sobre actividad económica).

═══════════════════════════════════════════════════════════════════════════════
ESTILO DE REDACCIÓN — TONO COMFENALCO (CRÍTICO, IMITAR EXACTO)
═══════════════════════════════════════════════════════════════════════════════
Comfenalco redacta este informe en estilo telegráfico profesional, denso en
datos y orientado a decisión operativa. Imita ESTE estilo en cada celda.

ESTRUCTURA DE CADA NOTA:
  [MARCADOR TEMPORAL opcional] [MUNICIPIO O VÍA en MAYÚSCULAS]: [hecho concreto
  con cifras y actores]. [Medidas: ... cuando aplique]. Fuente: [Nombre del medio].

MARCADORES TEMPORALES (en MAYÚSCULAS al inicio cuando aplique):
  • SEGUIMIENTO — hechos en curso desde hace días que merecen monitoreo continuo.
  • CONTINÚA — afectaciones vigentes, especialmente cierres viales que llevan días.
  • ALERTA — cuando hay declaración formal (alerta roja/naranja/amarilla).
  • (sin marcador) — hecho NUEVO de las últimas 72h.

EJEMPLOS REALES DEL CLIENTE (estudia y replica el tono y la densidad de datos):

— ORDEN PÚBLICO —
  «BRICEÑO: 86 familias desplazadas, ataques con explosivos, amenazas y multas
  a los campesinos, hacen que el municipio en este momento esté en alerta roja.
  Medidas: toque de queda y ley seca.»

  «YARUMAL: Ante el ataque sicarial en local comercial de Yarumal por un ajuste
  de cuentas con el Clan del Golfo, convocan de manera urgente a un consejo de
  seguridad, para tomar medidas de seguridad en el municipio.
  Fuente: Alerta Paisa y Teleantioquia.»

  «Oriente Antioqueño: Se mantiene una alerta por el incremento de homicidios
  derivados de enfrentamientos entre el Clan del Golfo y El Mesa, reportan en
  las últimas semanas 20 eventos aproximadamente. Fuente: Diarioriente.»

— VÍAS Y MOVILIDAD —
  «SEGUIMIENTO. Vía Nechí – Caucasia: Desde ayer presenta cierres intermitentes,
  específicamente en el sector de Colorado. Por parte del gremio de mineros.
  Fuente: Policía Nacional.»

  «CONTINÚA Cierre parcial – Ruta Nacional 2511 (Los Llanos – Tarazá).
  Tramo: PR 29+0530 al PR 31+000.
  Motivo: Obras de infraestructura y seguridad vial.
  Horarios de cierre: 09:00 a 15:00 / 20:00 a 05:00 del día siguiente.»

  «CONTINÚA Paso a un carril Dabeiba – Mutatá (sector Túnel La Llorona).
  Motivo: labores de mantenimiento y atención de puntos críticos.
  Fuente: Concesión Autopistas Urabá.»

  «CONTINÚA cierre total Vía Necoclí – Puerto Rey. Policía Nacional (DITRA) y
  reportes de INVÍAS confirman cierre por evento natural (daños en puentes y
  afectaciones por lluvias).»

— COMERCIO / SERVICIOS —
  «CONTINÚA. ITAGÜÍ: Más de 130 familias de la vereda El Porvenir, se quedaron
  sin servicio de agua potable luego de que vandalizaran la planta de tratamiento;
  la recuperación de la planta podría tardar cerca de ocho días.»

— SEDES / HOTELES / PARQUES —
  «CAÑAS GORDAS: Se presentará hoy interrupción de energía de 09:00 a 16:00 horas,
  por labores de mantenimiento. Zona urbana, rural y corregimientos.
  Fuente: La Noticia.»

REGLAS DE REDACCIÓN (basadas en los ejemplos anteriores):
  R1. Inicia cada nota con MUNICIPIO en MAYÚSCULAS (o nombre de vía), seguido
      de dos puntos. Si varios municipios de la subregión están afectados por
      el mismo hecho, agrúpalos en una sola nota.
  R2. Incluye SIEMPRE las cifras concretas que estén en el contexto: número de
      familias, víctimas, eventos, días, hectáreas, horarios. NO redondees, NO
      inventes. Si el dato no está, no lo pongas.
  R3. Identifica actores por nombre si aparecen en el contexto: «Clan del Golfo»,
      «El Mesa», «gremio de mineros», nombres de empresas, etc. NO inventes nombres.
  R4. Cierra con «Medidas: ...» cuando el contexto mencione decisiones tomadas
      (toque de queda, consejo de seguridad, ley seca, evacuación, militarización).
  R5. ⚠ FUENTES VERÍDICAS — REGLA CRÍTICA DE TRAZABILIDAD ⚠
      Cierra con «Fuente: <Nombre del medio>». El nombre DEBE coincidir EXACTAMENTE
      con el "📰 Medio:" que aparece junto al item que usaste. NO inventes nombres.
      NO uses "fuentes propias", "redes sociales", "según testigos" — cada hecho
      debe trazar a un medio listado en el contexto.
      Si combinas información de dos items en una sola nota: «Fuente: X y Y» listando ambos.
      Las URLs van en el campo "fuentes_*" del JSON — DEBEN ser las URLs exactas
      de los items que citaste con "Fuente: ..." en el texto. No omitas ninguna.
      Ejemplos del mapping nombre→dominio para que veas el patrón:
        "Alerta Paisa"          → alerta.com.co
        "Teleantioquia"         → teleantioquia.co
        "Diariente"             → diarioriente.com
        "Minuto30"              → minuto30.com
        "H13 Noticias"          → h13n.com
        "Conexión Sur"          → conexionsur.co
        "El Colombiano"         → elcolombiano.com
        "Policía Nacional"      → policia.gov.co
        "INVÍAS" / "INVÍAS Viajero" → invias.gov.co / invias-viajero.vercel.app
        "EPM"                   → epm.com.co
        "IDEAM"                 → ideam.gov.co
        "UNGRD"                 → ungrd.gov.co
        "Defensoría — Alertas Tempranas" → alertastempranas.defensoria.gov.co
        "Gobernación de Antioquia" → antioquia.gov.co
  R6. Para VÍAS: usa formato «Vía A – B» (guion largo –) seguido del tipo de
      afectación: «Cierre total», «Cierre parcial», «Paso a un carril»,
      «Paso restringido», «Cierres intermitentes». Añade Tramo (PR si lo hay),
      Motivo y Horarios si están en el contexto, en líneas separadas.
  R7. Para SERVICIOS PÚBLICOS: menciona barrio/vereda/zona, horario exacto
      («09:00 a 16:00»), duración estimada y motivo. Si afecta agua o energía
      en un municipio → propaga la nota TAMBIÉN a Sedes/Hoteles/Parques.
  R8. Si una subregión tiene varios hechos en la misma categoría, sepáralos con
      DOBLE salto de línea (\\n\\n) — NO los mezcles en un solo párrafo.
  R9. Longitud objetivo por nota: 2–5 líneas. Densas, no rellenadas.
  R10. Tono: telegráfico, neutro, profesional. NO uses adjetivos emocionales
       («terrible», «espantoso»). NO uses primera persona. NO uses futuro
       especulativo («podría suceder») salvo que el contexto lo diga textualmente.

═══════════════════════════════════════════════════════════════════════════════
CONTEXTO RECOPILADO
═══════════════════════════════════════════════════════════════════════════════
─── SEGURIDAD / ORDEN PÚBLICO ───
{seg}

─── VÍAS Y MOVILIDAD ───
{vias}

─── SERVICIOS PÚBLICOS Y COMERCIO ───
{svc}

─── SEDES, HOTELES Y PARQUES (Comfenalco Antioquia) ───
{sed}

═══════════════════════════════════════════════════════════════════════════════
INSTRUCCIONES DE PROCESAMIENTO
═══════════════════════════════════════════════════════════════════════════════
1. Asigna cada noticia/aviso a la subregión correspondiente. SI el item ya
   viene con etiqueta "📍 SUBREGIÓN_DETECTADA=X", úsala directamente — el
   sistema ya la pre-detectó por mención de municipio. NO la cambies salvo
   evidencia clara. Si NO trae esa etiqueta, asigna por el municipio mencionado.
2. Si un item trae "🎯 CRITICIDAD_SUGERIDA=X", úsala como punto de partida y
   AJÚSTALA solo si tu análisis del texto justifica otro nivel.
3. Si hay información: redacta SIGUIENDO EL ESTILO COMFENALCO de los ejemplos.
4. Sin información para la categoría/subregión → escribe exactamente: SIN NOVEDAD
5. Criticidad por categoría:
   GRAVES     = toque de queda / desplazamiento masivo / cierre total vía /
                ataque armado grave con víctimas / alerta roja
   RELEVANTES = alerta de seguridad / paso a un carril / corte prolongado
                de servicios / afectación parcial / alerta naranja
   PARCIALES  = SEGUIMIENTO / obra programada / alerta preventiva /
                corte corto con horario definido / alerta amarilla
   NINGUNA    = sin novedad
6. La criticidad_general de la subregión = la más alta entre sus 4 categorías.
7. En los campos "fuentes_*" devuelve SOLO URLs presentes en el contexto que SÍ
   usaste para esa categoría/subregión. Si no usaste ninguna → arreglo vacío [].

═══════════════════════════════════════════════════════════════════════════════
DEVUELVE SOLO ESTE JSON, sin texto antes ni después. Cada subregión debe tener
exactamente estas claves:
═══════════════════════════════════════════════════════════════════════════════
{{
  "BAJO CAUCA": {{
    "municipio_principal": "",
    "orden_publico": "", "orden_publico_crit": "NINGUNA", "fuentes_orden_publico": [],
    "vias": "", "vias_crit": "NINGUNA", "fuentes_vias": [],
    "comercio_servicios": "", "comercio_crit": "NINGUNA", "fuentes_comercio": [],
    "sedes_hoteles": "", "sedes_crit": "NINGUNA", "fuentes_sedes": [],
    "criticidad_general": "NINGUNA"
  }},
  "MAGDALENA MEDIO": {{ ... mismas claves ... }},
  "NORDESTE": {{ ... mismas claves ... }},
  "NORTE": {{ ... mismas claves ... }},
  "OCCIDENTE": {{ ... mismas claves ... }},
  "ORIENTE": {{ ... mismas claves ... }},
  "URABÁ": {{ ... mismas claves ... }},
  "SUROESTE": {{ ... mismas claves ... }},
  "ÁREA METROPOLITANA": {{ ... mismas claves ... }}
}}"""


# Cadena de modelos Groq — orden de prioridad por calidad → resiliencia ante rate-limits.
# Mezclamos proveedores (Meta, OpenAI OSS, Qwen) para diversificar cuotas.
# Lista actualizada: removidos los decommissioned (specdec, gemma2-9b-it, llama3-8b-8192).
GROQ_MODELOS = [
    "llama-3.3-70b-versatile",                       # Llama 3.3 70B — principal
    "meta-llama/llama-4-maverick-17b-128e-instruct", # Llama 4 Maverick (128 expertos)
    "meta-llama/llama-4-scout-17b-16e-instruct",     # Llama 4 Scout
    "openai/gpt-oss-120b",                           # GPT-OSS 120B (open weights)
    "qwen/qwen3-32b",                                # Qwen 3 — fuerte en multilingüe
    "openai/gpt-oss-20b",                            # GPT-OSS 20B (más rápido)
    "llama-3.1-8b-instant",                          # último recurso rápido
]

CRIT_ORDEN = {"NINGUNA": 0, "PARCIALES": 1, "RELEVANTES": 2, "GRAVES": 3}

def _peor_crit(*niveles) -> str:
    """Devuelve el nivel de criticidad más alto entre los dados."""
    n = max(niveles, key=lambda x: CRIT_ORDEN.get(x, 0))
    return n if n in CRIT_ORDEN else "NINGUNA"

def _validar_fuentes(fuentes_iter, urls_validas: set) -> list[str]:
    """Filtra URLs alucinadas: solo conserva las que vinieron en el contexto."""
    if not fuentes_iter:
        return []
    out = []
    for u in fuentes_iter:
        if isinstance(u, str) and u in urls_validas:
            out.append(u)
    # Dedupe preservando orden
    seen = set()
    return [u for u in out if not (u in seen or seen.add(u))]


def _texto_menciona_subregion(texto: str, subregion: str) -> bool:
    """
    Cross-mention validation RELAJADA: el contenido se conserva por defecto.
    Solo se rechaza si hay evidencia clara de CRUCE: el texto menciona un
    municipio de OTRA subregión y NO menciona ninguno de la subregión asignada.

    Casos aceptados:
      • Menciona la subregión literal ("Bajo Cauca", "Urabá")
      • Menciona un municipio de la subregión asignada
      • Menciona solo "Antioquia" en general (texto genérico válido)
      • Menciona "Antioquia" + un municipio del catálogo (la pre-detección lo manejará)
    Caso rechazado:
      • Menciona un municipio que pertenece a otra subregión Y NO menciona la actual
    """
    if not texto or texto.upper() == "SIN NOVEDAD":
        return True
    n = _normalize(texto)

    # 1. Menciona la subregión literal
    if _normalize(subregion) in n:
        return True

    # 2. Menciona un municipio de la subregión asignada
    municipios = SUBREGIONES.get(subregion, [])
    tokens = set(re.findall(r"[a-z]+", n))
    for m in municipios:
        mn = _normalize(m)
        if " " in mn:
            if mn in n:
                return True
        elif mn in tokens:
            return True

    # 3. Si solo menciona "antioquia" sin municipio específico → aceptar (texto genérico)
    sub_detectada = detectar_subregion(texto)
    if sub_detectada is None:
        return True

    # 4. Si detectó otra subregión → CRUCE confirmado, rechazar
    return False


def _esqueleto_vacio() -> dict:
    return {
        sub: {
            "municipio_principal": m[0],
            "orden_publico": "SIN NOVEDAD", "orden_publico_crit": "NINGUNA", "fuentes_orden_publico": [],
            "vias": "SIN NOVEDAD",           "vias_crit": "NINGUNA",          "fuentes_vias": [],
            "comercio_servicios": "SIN NOVEDAD", "comercio_crit": "NINGUNA",  "fuentes_comercio": [],
            "sedes_hoteles": "SIN NOVEDAD",  "sedes_crit": "NINGUNA",         "fuentes_sedes": [],
            "criticidad_general": "NINGUNA",
        }
        for sub, m in SUBREGIONES.items()
    }


_DEGRADACIONES: list[dict] = []

def analizar(groq_key: str, contexto: dict, fecha: str) -> dict:
    global _DEGRADACIONES
    _DEGRADACIONES = []  # reset por corrida
    client = Groq(api_key=groq_key)
    prompt = construir_prompt(
        contexto["texto_seg"], contexto["texto_vias"], contexto["texto_svc"],
        contexto["texto_sed"], fecha,
    )
    urls_validas = set(contexto["url_to_cat"].keys())
    ultimo_error = None

    for modelo in GROQ_MODELOS:
        try:
            resp = client.chat.completions.create(
                model=modelo,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=4000,
                temperature=0.1,
            )
            raw = resp.choices[0].message.content
            match = re.search(r"\{[\s\S]*\}", raw)
            if not match:
                continue
            parsed = json.loads(match.group())

            # Normalizar: rellenar claves faltantes + validar fuentes contra el contexto
            esqueleto = _esqueleto_vacio()
            for sub, base in esqueleto.items():
                ia = parsed.get(sub, {}) if isinstance(parsed, dict) else {}
                if not isinstance(ia, dict):
                    ia = {}

                base["municipio_principal"] = ia.get("municipio_principal") or base["municipio_principal"]

                for campo, fcampo, crit in [
                    ("orden_publico",      "fuentes_orden_publico", "orden_publico_crit"),
                    ("vias",               "fuentes_vias",          "vias_crit"),
                    ("comercio_servicios", "fuentes_comercio",      "comercio_crit"),
                    ("sedes_hoteles",      "fuentes_sedes",         "sedes_crit"),
                ]:
                    base[campo] = (ia.get(campo) or "SIN NOVEDAD").strip() or "SIN NOVEDAD"
                    base[crit]  = ia.get(crit) if ia.get(crit) in CRIT_ORDEN else "NINGUNA"
                    base[fcampo] = _validar_fuentes(ia.get(fcampo, []), urls_validas)

                    # Cross-mention validation: solo degrada si hay CRUCE real
                    # (texto que menciona otra subregión sin mencionar la actual).
                    if base[campo] != "SIN NOVEDAD" and not _texto_menciona_subregion(base[campo], sub):
                        _DEGRADACIONES.append({
                            "subregion": sub, "campo": campo,
                            "texto": base[campo][:160] + ("…" if len(base[campo]) > 160 else ""),
                            "sub_real": detectar_subregion(base[campo]) or "?",
                        })
                        base[campo]  = "SIN NOVEDAD"
                        base[crit]   = "NINGUNA"
                        base[fcampo] = []

                    # Si quedó SIN NOVEDAD, asegurar coherencia
                    if base[campo] == "SIN NOVEDAD":
                        base[fcampo] = []
                        base[crit]   = "NINGUNA"

                # Criticidad general = máxima entre las 4
                base["criticidad_general"] = _peor_crit(
                    base["orden_publico_crit"], base["vias_crit"],
                    base["comercio_crit"],     base["sedes_crit"],
                )

            st.info(f"🤖 Modelo usado: `{modelo}`")
            return esqueleto

        except Exception as e:
            ultimo_error = e
            msg = str(e)
            if "rate_limit" in msg or "429" in msg:
                st.warning(f"⏳ `{modelo}` sin cupo — probando siguiente…")
            elif "decommissioned" in msg or "deprecated" in msg:
                st.warning(f"🗑️ `{modelo}` descontinuado — probando siguiente…")
            else:
                st.warning(f"⚠️ `{modelo}` falló — probando siguiente…")
            continue

    if ultimo_error:
        st.error(f"❌ Todos los modelos fallaron. Último error: {ultimo_error}")
    return _esqueleto_vacio()


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — formato idéntico al original
# ══════════════════════════════════════════════════════════════════════════════
def construir_excel(datos: dict, fecha: str) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "REPORTE"

    def fill(hex6):
        return PatternFill("solid", fgColor=hex6)

    def borde_fino():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    # Fila 1: Título
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "INFORME DIARIO REGIONAL"
    t.font      = Font(bold=True, size=28, color="000000")
    t.fill      = fill("BBCE00")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 47

    # Fila 2: vacía
    ws.row_dimensions[2].height = 6

    # Fila 3: Encabezados
    ws["A3"].value         = datetime.strptime(fecha, "%d/%m/%Y")
    ws["A3"].number_format = "DD/MM/YYYY"
    ws["A3"].font          = Font(bold=True, size=24)
    ws["A3"].alignment     = Alignment(horizontal="center", vertical="center")

    hdrs = {
        "B3": "SUBREGIÓN",
        "C3": "MUNICIPIOS",
        "D3": "ORDEN PÚBLICO\n(Acciones de grupos criminales que pongan en riesgo parcial o totalmente la región o municipio)",
        "E3": "AFECTACIÓN DE VÍAS\n(Eventos naturales o de otra causa que afecten parcial o totalmente las vías)",
        "F3": "ESTADO COMERCIO O SERVICIOS\n(Situaciones ajenas al orden público que afecten parcial o totalmente la región o municipio)",
        "G3": "ESTADO DE SEDES, HOTELES Y PARQUES\n(Afectación de servicios públicos)",
    }
    for addr, val in hdrs.items():
        c = ws[addr]
        c.value     = val
        c.font      = Font(bold=True, size=14)
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        c.border    = borde_fino()
    ws.row_dimensions[3].height = 55

    # Filas de datos
    row = 4
    for sub, d in datos.items():
        def bg(crit_key):
            return COLORES_EXCEL.get(d.get(crit_key, "NINGUNA"), "92D050")

        cg  = d.get("criticidad_general", "NINGUNA")
        mun = d.get("municipio_principal", "")

        def texto_celda(campo, fuentes_campo):
            val = (d.get(campo) or "SIN NOVEDAD").strip() or "SIN NOVEDAD"
            if val == "SIN NOVEDAD":
                return "SIN NOVEDAD"
            fuentes = d.get(fuentes_campo) or []
            if fuentes:
                return val + "\n\nFuentes:\n" + "\n".join(fuentes)
            return val

        # Col A: indicador criticidad
        ca = ws.cell(row=row, column=1, value=LABELS_CRIT.get(cg, ""))
        ca.font      = Font(bold=True, size=10, color="000000")
        ca.fill      = fill(COLORES_EXCEL.get(cg, "92D050"))
        ca.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        ca.border    = borde_fino()

        # Col B: SUBREGIÓN
        cb = ws.cell(row=row, column=2, value=sub)
        cb.font      = Font(bold=True, size=12)
        cb.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cb.border    = borde_fino()

        # Col C: MUNICIPIO
        cc = ws.cell(row=row, column=3, value=mun)
        cc.font      = Font(bold=True, size=12)
        cc.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cc.border    = borde_fino()

        # Cols D-G
        contenidos = [
            (4, texto_celda("orden_publico",      "fuentes_orden_publico"), bg("orden_publico_crit")),
            (5, texto_celda("vias",               "fuentes_vias"),          bg("vias_crit")),
            (6, texto_celda("comercio_servicios", "fuentes_comercio"),      bg("comercio_crit")),
            (7, texto_celda("sedes_hoteles",      "fuentes_sedes"),         bg("sedes_crit")),
        ]
        for col, val, color in contenidos:
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(size=15)
            c.fill      = fill(color)
            c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            c.border    = borde_fino()

        max_len = max(
            len(d.get("orden_publico", "") or ""),
            len(d.get("vias", "") or ""),
            len(d.get("comercio_servicios", "") or ""),
            len(d.get("sedes_hoteles", "") or ""),
        )
        ws.row_dimensions[row].height = max(80, min(max_len * 0.6, 400))
        row += 1

    # Leyenda
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    lc = ws[f"A{row}"]
    lc.value     = "NIVEL CRITICIDAD ORDEN PÚBLICO A LA FECHA"
    lc.font      = Font(bold=True, size=11)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 42

    leyenda = [
        ("NO HAY\nAFECTACIONES",    "92D050", "BAJA"),
        ("AFECTACIONES\nPARCIALES", "FFFF00", "MEDIA"),
        ("AFECTACIONES\nRELEVANTES","FFC000", "ALTA"),
        ("AFECTACIONES\nGRAVES",    "FF0000", "MUY ALTA"),
    ]
    for i, (label, color, nivel) in enumerate(leyenda):
        r = row + 1 + i
        ws.merge_cells(f"A{r}:B{r}")
        c = ws[f"A{r}"]
        c.value     = label
        c.fill      = fill(color)
        c.font      = Font(bold=True, size=11)
        c.border    = borde_fino()
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        ws[f"C{r}"].value = nivel
        ws[f"C{r}"].font  = Font(bold=True, size=11)
        ws.row_dimensions[r].height = 50

    contactos = {
        f"D{row}": "INVÍAS / X: @numeral767 | Mintransporte / X: @MintransporteCo | Policía Vías / X: @PoliciaDeTransito",
        f"E{row}": "Urabá: 300895777 | Bogotá: 3176460735 | Pacífico: 300895777 | Cauya–La Pintada: 3142220000 | Cañasgordas–Mutatá: 3142220000",
        f"F{row}": "invias-viajero.vercel.app | policia.gov.co/estado-de-las-vias | @AutopistasUraba | @AutopistasCafé | @Covipacifico | @ConcesionL",
        f"G{row}": "Protocolo Informativo Desplazamientos\n\n1. CIMCA: 3104382572 - 6045108558\n2. Antes de viajar: Consultar INVÍAS 018005191656\n3. En ruta: WhatsApp 317 646 0735\n4. Emergencia: 123 Policía / 3203017300 INVIAS",
    }
    for addr, txt in contactos.items():
        c = ws[addr]
        c.value     = txt
        c.font      = Font(size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    anchos = {
        "A": 36.5, "B": 18.9, "C": 20.0,
        "D": 88.1, "E": 87.9, "F": 79.4, "G": 72.7,
    }
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
#  UI
# ══════════════════════════════════════════════════════════════════════════════
hoy = datetime.now()

st.logo("https://i.imgur.com/RFdkaOo.png")

# Logo oficial de Comfenalco
_LOGO_URL = "https://i.imgur.com/RFdkaOo.png"

_fecha_completa = fecha_espanol(hoy)
_dia_num = f"{hoy.day:02d}"
_mes_anio = _fecha_completa.split(' DE ')[1] if ' DE ' in _fecha_completa else _fecha_completa

st.markdown(f"""
<div class="hero">
    <div class="hero-row">
        <div class="hero-left">
            <div class="hero-logo">
                <img src="{_LOGO_URL}" alt="Comfenalco">
            </div>
            <div class="hero-text">
                <div class="hero-tag"><span class="dot"></span> Comfenalco Antioquia</div>
                <h1 class="hero-title">Informe <em>Diario Regional</em></h1>
            </div>
        </div>
        <div class="hero-date">
            <span class="day">{_dia_num}</span>
            {_mes_anio}
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

izq, der = st.columns([1, 2.3], gap="large")

with izq:
    st.markdown('<p class="panel-title">Configuración</p>', unsafe_allow_html=True)

    try:
        groq_key = st.secrets["GROQ_API_KEY"]
        st.success("Groq configurado")
    except Exception:
        groq_key = st.text_input("Groq API Key", type="password",
                                  placeholder="gsk_...",
                                  help="Gratis en console.groq.com")

    try:
        tavily_key = st.secrets["TAVILY_API_KEY"]
        st.success("Tavily configurado")
    except Exception:
        tavily_key = st.text_input("Tavily API Key", type="password",
                                    placeholder="tvly-...",
                                    help="Gratis en app.tavily.com")

    fecha_informe = st.date_input("Fecha del informe", value=hoy)

    st.markdown("---")
    st.markdown('<p class="panel-title">Subregiones</p>', unsafe_allow_html=True)
    seleccion = {sub: st.checkbox(sub, value=True, key=f"chk_{sub}") for sub in SUBREGIONES}

    n_activas = sum(seleccion.values())
    st.markdown(
        f'<div style="font-size:0.78rem;color:var(--ink-4);margin-top:0.8rem;'
        f'padding:0.4rem 0;letter-spacing:0.2px">'
        f'<b style="color:var(--brand);font-weight:600">{n_activas}</b> '
        f'de {len(SUBREGIONES)} subregiones activas</div>',
        unsafe_allow_html=True,
    )

with der:
    generar = st.button("Generar informe", use_container_width=True)

    if generar:
        if not groq_key or not tavily_key:
            st.error("❌ Necesitas ambas API Keys (Groq y Tavily).")
            st.stop()

        fecha_str = fecha_informe.strftime("%d/%m/%Y")
        barra  = st.progress(0)
        estado = st.empty()

        estado.markdown("**Buscando en medios colombianos…** (últimos 7 días, mención de Antioquia)")
        barra.progress(15)
        contexto = obtener_noticias(tavily_key)
        barra.progress(45)

        estado.markdown("**Analizando y redactando con IA…** Groq · Llama 3.3 70B")
        datos = analizar(groq_key, contexto, fecha_str)
        datos = {s: datos[s] for s in SUBREGIONES if seleccion.get(s) and s in datos}
        barra.progress(85)

        estado.markdown("**Generando Excel…**")
        excel_buf = construir_excel(datos, fecha_str)
        barra.progress(100)
        estado.markdown("**Informe listo**")

        # ── Métricas ──
        st.markdown("---")
        conteo = {"GRAVES": 0, "RELEVANTES": 0, "PARCIALES": 0, "NINGUNA": 0}
        for d in datos.values():
            cg = d.get("criticidad_general", "NINGUNA")
            conteo[cg] = conteo.get(cg, 0) + 1

        _stat_meta = [
            ("GRAVES",     "Graves",      "graves",     "var(--crit-r)"),
            ("RELEVANTES", "Relevantes",  "relevantes", "var(--crit-o)"),
            ("PARCIALES",  "Parciales",   "parciales",  "var(--crit-y)"),
            ("NINGUNA",    "Sin novedad", "ninguna",    "var(--crit-g)"),
        ]
        _cards = "".join(
            f'<div class="stat-card" style="--sc:{color}">'
            f'<div class="stat-num">{conteo[k]}</div>'
            f'<div class="stat-lab"><span class="sdot sdot-{slug}"></span>{lab}</div>'
            f'</div>'
            for k, lab, slug, color in _stat_meta
        )
        st.markdown(f'<div class="stat-grid">{_cards}</div>', unsafe_allow_html=True)

        # ── Avisos oficiales vigentes (modelo bitemporal, distintos a noticias 72h) ──
        cierres_activos   = contexto.get("cierres_activos", [])
        avisos_servicios  = contexto.get("avisos_servicios", [])
        total_avisos = len(cierres_activos) + len(avisos_servicios)

        def _render_aviso(c: dict, badge_label: str, badge_color: str):
            marca = "🔁 INDEFINIDO" if c["indefinido"] else f"⏳ Vigente hasta {c['fecha_fin']}"
            st.markdown(
                f"""
                <div class="news-card" style="border-left-color:{badge_color};">
                    <div class="news-meta">
                        <span style="background:{badge_color};color:white;padding:2px 8px;border-radius:6px;font-weight:700;letter-spacing:0.5px;">{badge_label}</span>
                        &nbsp;·&nbsp; {marca}
                    </div>
                    <div class="news-snippet">{c['texto']}</div>
                    <a href="{c['fuente']}" target="_blank" class="source-tag">↗ Fuente oficial</a>
                </div>
                """,
                unsafe_allow_html=True,
            )

        if total_avisos > 0:
            with st.expander(f"🚨 Avisos oficiales vigentes hoy ({total_avisos})", expanded=True):
                st.caption("Datos estructurados de fuentes oficiales. Vigencia por ventana de validez, NO por fecha de publicación.")
                tab_v, tab_s = st.tabs([
                    f"🛣️ Cierres viales ({len(cierres_activos)})",
                    f"⚡ Servicios públicos ({len(avisos_servicios)})",
                ])
                with tab_v:
                    if cierres_activos:
                        for c in cierres_activos:
                            _render_aviso(c, "OFICIAL · CIERRE VIAL", "#ef4444")
                    else:
                        st.info("Sin cierres viales vigentes en fuentes oficiales hoy.")
                with tab_s:
                    if avisos_servicios:
                        for c in avisos_servicios:
                            _render_aviso(c, "OFICIAL · SERVICIO", "#0ea5e9")
                    else:
                        st.info("Sin avisos de servicios vigentes en fuentes oficiales hoy.")

        # ── Noticias fuente (cards) ──
        n_total = len(contexto["indexadas"])
        with st.expander(f"🔎 Ver noticias fuente ({n_total} verificadas, últimas 72h)"):
            t1, t2, t3, t4 = st.tabs(["🚨 Seguridad", "🛣️ Vías", "🏪 Servicios", "🏨 Sedes/Hoteles"])
            cat_label = {"seg": t1, "vias": t2, "svc": t3, "sed": t4}
            por_cat = {"seg": [], "vias": [], "svc": [], "sed": []}
            for it in contexto["indexadas"]:
                por_cat[contexto["url_to_cat"].get(it["url"], "seg")].append(it)

            for cat, tab in cat_label.items():
                with tab:
                    items = por_cat[cat]
                    if not items:
                        st.info("Sin noticias recientes en medios colombianos para esta categoría.")
                        continue
                    for it in items:
                        tier_label, tier_color = TIER_BADGE.get(it.get("tier", 1), ("REGIONAL", "#64748b"))
                        st.markdown(
                            f"""
                            <div class="news-card">
                                <div class="news-meta">
                                    <span style="background:{tier_color};color:white;padding:2px 8px;border-radius:6px;font-weight:700;letter-spacing:0.5px;">{tier_label}</span>
                                    &nbsp;·&nbsp; 📅 {it['fecha']} &nbsp;·&nbsp; score {it['score']:.2f}
                                </div>
                                <div class="news-title">{it['titulo']}</div>
                                <div class="news-snippet">{it['resumen'][:280]}…</div>
                                <a href="{it['url']}" target="_blank" class="source-tag">↗ Abrir fuente</a>
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )

        # ── Diagnóstico (transparencia del pipeline) ──
        total_oficiales = len(cierres_activos) + len(avisos_servicios)
        subs_con_contenido = sum(
            1 for d in datos.values()
            if any(d.get(k, "SIN NOVEDAD") != "SIN NOVEDAD" for k in
                   ["orden_publico", "vias", "comercio_servicios", "sedes_hoteles"])
        )
        n_degradaciones = len(_DEGRADACIONES)

        with st.expander(f"🔬 Diagnóstico del pipeline (transparencia)", expanded=(subs_con_contenido <= 1)):
            dc1, dc2, dc3, dc4 = st.columns(4)
            dc1.metric("Noticias procesadas", n_total)
            dc2.metric("Avisos oficiales", total_oficiales)
            dc3.metric("Subregiones con contenido", f"{subs_con_contenido}/{len(datos)}")
            dc4.metric("Degradaciones (cruce)", n_degradaciones)

            # Stats de búsqueda Tavily — cuántas trajo, cuántas cayeron y por qué
            if _SEARCH_STATS:
                st.markdown("**Búsqueda Tavily por categoría** (raw → descartes → kept):")
                cat_nombre = {"seg": "🚨 Seguridad", "vias": "🛣️ Vías", "svc": "🏪 Servicios", "sed": "🏨 Sedes/Hoteles"}
                for cat, s in _SEARCH_STATS.items():
                    nombre = cat_nombre.get(cat, cat)
                    st.markdown(
                        f"- **{nombre}** — Tavily devolvió **{s['raw']}** resultados "
                        f"({s.get('sub_queries', 1)} sub-queries) · "
                        f"descartes: viejos **{s.get('too_old', 0)}**, "
                        f"sin-Antioquia **{s.get('no_geo', 0)}**, "
                        f"ruido/farándula **{s.get('ruido', 0)}**, "
                        f"duplicados **{s.get('dup', 0)}**, "
                        f"basura **{s.get('no_basics', 0)}** · "
                        f"~ fecha-aprox **{s.get('fecha_aprox', 0)}** (aceptados con penalización) · "
                        f"✅ **{s.get('kept', 0)}** kept"
                    )

            if n_degradaciones > 0:
                st.markdown("**Contenido degradado por cross-mention validation** (la IA puso contenido en una subregión pero el texto mencionaba otra):")
                for d in _DEGRADACIONES:
                    st.markdown(
                        f"- ⚠ **{d['subregion']} · {d['campo']}** → mejor candidato: `{d['sub_real']}`<br>"
                        f"  <span style='color:#64748b;font-size:0.85rem'>«{d['texto']}»</span>",
                        unsafe_allow_html=True,
                    )

            if subs_con_contenido == 0:
                st.warning(
                    "⚠ Ninguna subregión tiene contenido. Verifica:\n"
                    f"• La fecha del informe es {fecha_str} — ¿es correcta?\n"
                    "• ¿Hay noticias en el expander de fuentes? Si no, posiblemente el día está tranquilo.\n"
                    "• Revisa que las API keys de Tavily/Groq estén activas."
                )

        # ── Resumen por subregión (cards con borde coloreado) ──
        st.markdown('<div class="section-h">Resumen de Novedades por Subregión</div>', unsafe_allow_html=True)
        crit_class = {
            "GRAVES":     "crit-graves",
            "RELEVANTES": "crit-relevantes",
            "PARCIALES":  "crit-parciales",
            "NINGUNA":    "crit-ninguna",
        }
        for sub, d in datos.items():
            crit    = d.get("criticidad_general", "NINGUNA")
            mun     = d.get("municipio_principal", "")

            st.markdown(f'<div class="{crit_class.get(crit, "")}">', unsafe_allow_html=True)
            with st.expander(f"{sub}   ·   {mun}   ·   {LABELS_CRIT.get(crit, '')}"):
                c1, c2 = st.columns(2)
                with c1:
                    oc = d.get('orden_publico_crit', 'NINGUNA')
                    st.markdown(f"{dot(oc)}**Orden Público**", unsafe_allow_html=True)
                    st.caption(d.get("orden_publico", "SIN NOVEDAD"))
                    if d.get("fuentes_orden_publico"):
                        for u in d["fuentes_orden_publico"]:
                            st.markdown(f'<a href="{u}" target="_blank" class="source-tag">↗ {u[:50]}…</a>', unsafe_allow_html=True)

                    vc = d.get('vias_crit', 'NINGUNA')
                    st.markdown(f"{dot(vc)}**Vías**", unsafe_allow_html=True)
                    st.caption(d.get("vias", "SIN NOVEDAD"))
                    if d.get("fuentes_vias"):
                        for u in d["fuentes_vias"]:
                            st.markdown(f'<a href="{u}" target="_blank" class="source-tag">↗ {u[:50]}…</a>', unsafe_allow_html=True)
                with c2:
                    cc = d.get('comercio_crit', 'NINGUNA')
                    st.markdown(f"{dot(cc)}**Comercio / Servicios**", unsafe_allow_html=True)
                    st.caption(d.get("comercio_servicios", "SIN NOVEDAD"))
                    if d.get("fuentes_comercio"):
                        for u in d["fuentes_comercio"]:
                            st.markdown(f'<a href="{u}" target="_blank" class="source-tag">↗ {u[:50]}…</a>', unsafe_allow_html=True)

                    sc = d.get('sedes_crit', 'NINGUNA')
                    st.markdown(f"{dot(sc)}**Sedes / Hoteles**", unsafe_allow_html=True)
                    st.caption(d.get("sedes_hoteles", "SIN NOVEDAD"))
                    if d.get("fuentes_sedes"):
                        for u in d["fuentes_sedes"]:
                            st.markdown(f'<a href="{u}" target="_blank" class="source-tag">↗ {u[:50]}…</a>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

        # ── Descarga ──
        st.markdown("---")
        nombre = f"Informe_Diario_Regional_{fecha_informe.strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="📥 DESCARGAR EXCEL",
            data=excel_buf, file_name=nombre,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption(f"Listo para revisar y convertir a PDF: `{nombre}`")

    else:
        st.markdown("""
        <div class="placeholder">
            <div class="placeholder-glyph">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round">
                    <circle cx="11" cy="11" r="8"/>
                    <line x1="21" y1="21" x2="16.65" y2="16.65"/>
                </svg>
            </div>
            <h2>Listo para generar</h2>
            <p>Pulsa <b style="color:var(--brand);font-weight:600">Generar informe</b> en el panel lateral para iniciar el rastreo.</p>
        </div>
        """, unsafe_allow_html=True)

# ── Footer ───────────────────────────────────────────────────────────────────
st.markdown("""
<div class="brand-foot">
    <b>Comfenalco Antioquia</b> · Informe Diario Regional
</div>
""", unsafe_allow_html=True)
